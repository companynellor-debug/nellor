from fastapi import FastAPI, APIRouter, UploadFile, File, Form, HTTPException
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os, logging, uuid, json, io, csv, re, traceback
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime, timezone
import httpx
from PyPDF2 import PdfReader
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'test_database')]

# Supabase config
SUPABASE_URL = os.environ.get('SUPABASE_URL', '')
SUPABASE_SERVICE_KEY = os.environ.get('SUPABASE_SERVICE_ROLE_KEY', '')
GROQ_API_KEY = os.environ.get('GROQ_API_KEY', '')

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ──────── Groq helpers ────────

GROQ_ENDPOINT = "https://api.groq.com/openai/v1/chat/completions"
GROQ_MODEL = "llama-3.3-70b-versatile"

async def call_groq(prompt: str, max_retries: int = 3) -> str:
    """Call Groq API with retry logic."""
    headers = {
        "Authorization": f"Bearer {GROQ_API_KEY}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": GROQ_MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.1,
        "max_tokens": 8000,
    }
    for attempt in range(max_retries):
        try:
            async with httpx.AsyncClient(timeout=60) as client_http:
                resp = await client_http.post(GROQ_ENDPOINT, headers=headers, json=payload)
                if resp.status_code == 200:
                    data = resp.json()
                    return data["choices"][0]["message"]["content"]
                else:
                    logger.warning(f"Groq attempt {attempt+1} failed: {resp.status_code} {resp.text[:200]}")
        except Exception as e:
            logger.warning(f"Groq attempt {attempt+1} error: {e}")
        if attempt < max_retries - 1:
            import asyncio
            await asyncio.sleep(2)
    raise HTTPException(status_code=502, detail="Falha ao processar com IA após 3 tentativas")


def parse_groq_json(text: str) -> list:
    """Extract JSON array from Groq response."""
    text = text.strip()
    # Try to find JSON array
    match = re.search(r'\[.*\]', text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group())
        except json.JSONDecodeError:
            pass
    # Try parsing the whole response
    try:
        result = json.loads(text)
        if isinstance(result, list):
            return result
        if isinstance(result, dict) and "products" in result:
            return result["products"]
        return [result]
    except json.JSONDecodeError:
        return []


def compute_confidence(parsed: dict) -> str:
    """Compute confidence level based on filled fields."""
    fields = ["name", "price", "description", "category", "brand", "min_order", "variations"]
    filled = sum(1 for f in fields if parsed.get(f))
    if filled >= 5:
        return "alta"
    if filled >= 3:
        return "media"
    return "baixa"


# ──────── Supabase helpers ────────

async def supabase_insert_product(product_data: dict) -> dict:
    """Insert a product into Supabase products table."""
    headers = {
        "apikey": SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }
    async with httpx.AsyncClient(timeout=15) as c:
        resp = await c.post(f"{SUPABASE_URL}/rest/v1/products", headers=headers, json=product_data)
        if resp.status_code in (200, 201):
            data = resp.json()
            return data[0] if isinstance(data, list) else data
        else:
            logger.error(f"Supabase insert error: {resp.status_code} {resp.text[:300]}")
            raise HTTPException(status_code=500, detail=f"Erro ao criar produto: {resp.text[:200]}")


CATALOG_PROMPT_PDF = """Analise este catálogo de produtos e extraia todos os produtos encontrados.
Para cada produto retorne um JSON com os campos:
- name: nome do produto
- description: descrição completa
- price: preço (número apenas, sem R$)
- min_order: quantidade mínima de pedido se mencionada
- brand: marca se mencionada
- variations: array de variações encontradas (cores, tamanhos, modelos)
- category: categoria sugerida
- sale_unit: unidade de venda (unit, box, bale, kit, pair)

Retorne APENAS um array JSON válido sem texto adicional.
Texto do catálogo:
"""

CATALOG_PROMPT_URL = """Analise o conteúdo desta página de produto/catálogo e extraia todos os produtos encontrados.
Para cada produto retorne um JSON com os campos:
- name: nome do produto
- description: descrição completa
- price: preço (número apenas, sem R$)
- min_order: quantidade mínima de pedido se mencionada
- brand: marca se mencionada
- variations: array de variações encontradas (cores, tamanhos, modelos)
- category: categoria sugerida
- sale_unit: unidade de venda (unit, box, bale, kit, pair)
- image_url: URL da imagem do produto se encontrada

Retorne APENAS um array JSON válido sem texto adicional.
Conteúdo da página:
"""

COLUMN_MAPPING_PROMPT = """Analise os nomes das colunas desta planilha e mapeie cada uma para o campo correto do sistema.

Campos do sistema:
- name: nome do produto
- description: descrição
- price: preço
- min_order: quantidade mínima
- brand: marca
- category: categoria
- variations: variações (cores, tamanhos)
- sale_unit: unidade de venda

Colunas encontradas: {columns}

Retorne APENAS um JSON com o mapeamento no formato:
{{"coluna_original": "campo_sistema", ...}}
Se uma coluna não corresponder a nenhum campo, mapeie como null.
"""


# ──────── API routes ────────

@api_router.get("/")
async def root():
    return {"message": "Nellor API"}


@api_router.post("/catalog/import/pdf")
async def import_pdf(
    file: UploadFile = File(...),
    supplier_id: str = Form(...)
):
    """Process PDF catalog import."""
    if not file.filename or not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Apenas arquivos PDF são aceitos")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Arquivo muito grande (máximo 10MB)")

    import_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()

    # Create import record
    import_doc = {
        "id": import_id,
        "supplier_id": supplier_id,
        "import_type": "pdf",
        "source_file_url": file.filename,
        "status": "processing",
        "products_found": 0,
        "products_imported": 0,
        "error_message": None,
        "created_at": now,
        "completed_at": None,
    }
    await db.catalog_imports.insert_one(import_doc)

    try:
        # Extract text from PDF
        reader = PdfReader(io.BytesIO(content))
        text = ""
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"

        if not text.strip():
            await db.catalog_imports.update_one(
                {"id": import_id},
                {"$set": {"status": "failed", "error_message": "Não foi possível extrair texto do PDF"}}
            )
            return {"import_id": import_id, "status": "failed", "error": "PDF sem texto extraível"}

        # Truncate to avoid token limits
        text = text[:12000]

        # Call Groq
        groq_response = await call_groq(CATALOG_PROMPT_PDF + text)
        products = parse_groq_json(groq_response)

        if not products:
            await db.catalog_imports.update_one(
                {"id": import_id},
                {"$set": {"status": "failed", "error_message": "Nenhum produto encontrado no PDF"}}
            )
            return {"import_id": import_id, "status": "failed", "error": "Nenhum produto encontrado"}

        # Save extracted products
        import_products = []
        for p in products:
            prod_doc = {
                "id": str(uuid.uuid4()),
                "import_id": import_id,
                "raw_data": p,
                "parsed_data": {
                    "name": p.get("name", ""),
                    "description": p.get("description", ""),
                    "price": p.get("price"),
                    "min_order": p.get("min_order"),
                    "brand": p.get("brand"),
                    "variations": p.get("variations", []),
                    "category": p.get("category", ""),
                    "sale_unit": p.get("sale_unit", "unit"),
                    "confidence": compute_confidence(p),
                },
                "status": "pending",
                "product_id": None,
                "created_at": now,
            }
            import_products.append(prod_doc)

        if import_products:
            await db.catalog_import_products.insert_many(import_products)

        await db.catalog_imports.update_one(
            {"id": import_id},
            {"$set": {"status": "review", "products_found": len(import_products)}}
        )

        return {
            "import_id": import_id,
            "status": "review",
            "products_found": len(import_products),
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"PDF import error: {traceback.format_exc()}")
        await db.catalog_imports.update_one(
            {"id": import_id},
            {"$set": {"status": "failed", "error_message": str(e)[:500]}}
        )
        raise HTTPException(status_code=500, detail=f"Erro ao processar PDF: {str(e)[:200]}")


@api_router.post("/catalog/import/spreadsheet")
async def import_spreadsheet(
    file: UploadFile = File(...),
    supplier_id: str = Form(...)
):
    """Process spreadsheet (xlsx, xls, csv) catalog import."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="Arquivo inválido")

    ext = file.filename.lower().rsplit('.', 1)[-1] if '.' in file.filename else ''
    if ext not in ('xlsx', 'xls', 'csv'):
        raise HTTPException(status_code=400, detail="Apenas .xlsx, .xls ou .csv")

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Arquivo muito grande (máximo 10MB)")

    import_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()

    import_doc = {
        "id": import_id,
        "supplier_id": supplier_id,
        "import_type": "spreadsheet",
        "source_file_url": file.filename,
        "status": "processing",
        "products_found": 0,
        "products_imported": 0,
        "error_message": None,
        "created_at": now,
        "completed_at": None,
    }
    await db.catalog_imports.insert_one(import_doc)

    try:
        rows = []
        headers_list = []

        if ext == 'csv':
            text = content.decode('utf-8', errors='replace')
            reader_csv = csv.DictReader(io.StringIO(text))
            headers_list = reader_csv.fieldnames or []
            for row in reader_csv:
                rows.append(dict(row))
        else:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                raise HTTPException(status_code=400, detail="Planilha vazia")
            all_rows = list(ws.iter_rows(values_only=True))
            if len(all_rows) < 2:
                raise HTTPException(status_code=400, detail="Planilha sem dados suficientes")
            headers_list = [str(h) if h else f"col_{i}" for i, h in enumerate(all_rows[0])]
            for row_vals in all_rows[1:]:
                row_dict = {}
                for i, val in enumerate(row_vals):
                    if i < len(headers_list):
                        row_dict[headers_list[i]] = val
                if any(v is not None and str(v).strip() for v in row_dict.values()):
                    rows.append(row_dict)

        if not rows:
            await db.catalog_imports.update_one(
                {"id": import_id},
                {"$set": {"status": "failed", "error_message": "Planilha vazia ou sem dados"}}
            )
            return {"import_id": import_id, "status": "failed", "error": "Planilha vazia"}

        # Use Groq for column mapping
        mapping = {}
        try:
            mapping_response = await call_groq(COLUMN_MAPPING_PROMPT.format(columns=json.dumps(headers_list, ensure_ascii=False)))
            mapping_match = re.search(r'\{.*\}', mapping_response, re.DOTALL)
            if mapping_match:
                mapping = json.loads(mapping_match.group())
        except Exception as e:
            logger.warning(f"Column mapping failed: {e}")

        # Fallback: simple keyword matching
        if not mapping:
            keyword_map = {
                "name": ["nome", "name", "produto", "product", "título", "titulo", "item"],
                "description": ["descrição", "descricao", "description", "desc", "detalhes"],
                "price": ["preço", "preco", "price", "valor", "custo", "value"],
                "min_order": ["mínimo", "minimo", "min", "qtd_min", "minimum"],
                "brand": ["marca", "brand", "fabricante"],
                "category": ["categoria", "category", "tipo", "type"],
                "variations": ["variações", "variacoes", "variations", "cores", "tamanhos"],
                "sale_unit": ["unidade", "unit", "und", "un"],
            }
            for header in headers_list:
                h_lower = header.lower().strip()
                for field, keywords in keyword_map.items():
                    if any(k in h_lower for k in keywords):
                        mapping[header] = field
                        break

        # Map rows to products
        import_products = []
        for row in rows[:200]:  # Limit to 200 products
            parsed = {}
            raw = {}
            for header, value in row.items():
                raw[header] = str(value) if value is not None else ""
                field = mapping.get(header)
                if field and value is not None:
                    str_val = str(value).strip()
                    if field == "price":
                        try:
                            clean = re.sub(r'[^\d.,]', '', str_val).replace(',', '.')
                            parsed["price"] = float(clean) if clean else None
                        except ValueError:
                            parsed["price"] = None
                    elif field == "min_order":
                        try:
                            parsed["min_order"] = int(re.sub(r'[^\d]', '', str_val) or "0")
                        except ValueError:
                            parsed["min_order"] = None
                    elif field == "variations":
                        parsed["variations"] = [v.strip() for v in str_val.split(',') if v.strip()]
                    else:
                        parsed[field] = str_val

            if not parsed.get("name"):
                continue

            parsed.setdefault("description", "")
            parsed.setdefault("category", "")
            parsed.setdefault("sale_unit", "unit")
            parsed["confidence"] = compute_confidence(parsed)

            prod_doc = {
                "id": str(uuid.uuid4()),
                "import_id": import_id,
                "raw_data": raw,
                "parsed_data": parsed,
                "status": "pending",
                "product_id": None,
                "created_at": now,
            }
            import_products.append(prod_doc)

        if import_products:
            await db.catalog_import_products.insert_many(import_products)

        await db.catalog_imports.update_one(
            {"id": import_id},
            {"$set": {
                "status": "review",
                "products_found": len(import_products),
            }}
        )

        return {
            "import_id": import_id,
            "status": "review",
            "products_found": len(import_products),
            "column_mapping": mapping,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Spreadsheet import error: {traceback.format_exc()}")
        await db.catalog_imports.update_one(
            {"id": import_id},
            {"$set": {"status": "failed", "error_message": str(e)[:500]}}
        )
        raise HTTPException(status_code=500, detail=f"Erro ao processar planilha: {str(e)[:200]}")


@api_router.post("/catalog/import/url")
async def import_url(
    supplier_id: str = Form(...),
    url: str = Form(...)
):
    """Process URL catalog import."""
    if not url or not url.startswith(('http://', 'https://')):
        raise HTTPException(status_code=400, detail="URL inválida")

    import_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()

    import_doc = {
        "id": import_id,
        "supplier_id": supplier_id,
        "import_type": "url",
        "source_url": url,
        "status": "processing",
        "products_found": 0,
        "products_imported": 0,
        "error_message": None,
        "created_at": now,
        "completed_at": None,
    }
    await db.catalog_imports.insert_one(import_doc)

    try:
        # Fetch URL content
        async with httpx.AsyncClient(timeout=15, follow_redirects=True) as c:
            headers_req = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            }
            resp = await c.get(url, headers=headers_req)

        if resp.status_code != 200:
            await db.catalog_imports.update_one(
                {"id": import_id},
                {"$set": {"status": "failed", "error_message": f"Não conseguimos acessar este site (HTTP {resp.status_code})"}}
            )
            return {"import_id": import_id, "status": "failed", "error": "Não conseguimos acessar este site automaticamente. Tente usar o PDF ou planilha."}

        html = resp.text
        soup = BeautifulSoup(html, 'html.parser')

        # Remove scripts, styles
        for tag in soup(['script', 'style', 'noscript', 'iframe', 'svg']):
            tag.decompose()

        # Extract images
        images = []
        for img in soup.find_all('img', src=True):
            src = img['src']
            if src.startswith('//'):
                src = 'https:' + src
            elif src.startswith('/'):
                from urllib.parse import urljoin
                src = urljoin(url, src)
            if src.startswith('http') and any(ext in src.lower() for ext in ['.jpg', '.jpeg', '.png', '.webp']):
                images.append(src)

        text = soup.get_text(separator='\n', strip=True)
        text = re.sub(r'\n{3,}', '\n\n', text)
        text = text[:12000]

        # Append image URLs to context
        img_context = ""
        if images[:20]:
            img_context = "\n\nImagens encontradas:\n" + "\n".join(images[:20])

        groq_response = await call_groq(CATALOG_PROMPT_URL + text + img_context)
        products = parse_groq_json(groq_response)

        if not products:
            await db.catalog_imports.update_one(
                {"id": import_id},
                {"$set": {"status": "failed", "error_message": "Nenhum produto encontrado na página"}}
            )
            return {"import_id": import_id, "status": "failed", "error": "Nenhum produto encontrado"}

        import_products = []
        for p in products:
            prod_doc = {
                "id": str(uuid.uuid4()),
                "import_id": import_id,
                "raw_data": p,
                "parsed_data": {
                    "name": p.get("name", ""),
                    "description": p.get("description", ""),
                    "price": p.get("price"),
                    "min_order": p.get("min_order"),
                    "brand": p.get("brand"),
                    "variations": p.get("variations", []),
                    "category": p.get("category", ""),
                    "sale_unit": p.get("sale_unit", "unit"),
                    "image_url": p.get("image_url"),
                    "confidence": compute_confidence(p),
                },
                "status": "pending",
                "product_id": None,
                "created_at": now,
            }
            import_products.append(prod_doc)

        if import_products:
            await db.catalog_import_products.insert_many(import_products)

        await db.catalog_imports.update_one(
            {"id": import_id},
            {"$set": {"status": "review", "products_found": len(import_products)}}
        )

        return {
            "import_id": import_id,
            "status": "review",
            "products_found": len(import_products),
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"URL import error: {traceback.format_exc()}")
        await db.catalog_imports.update_one(
            {"id": import_id},
            {"$set": {"status": "failed", "error_message": str(e)[:500]}}
        )
        raise HTTPException(status_code=500, detail=f"Erro: {str(e)[:200]}")


@api_router.get("/catalog/imports/{supplier_id}")
async def get_imports(supplier_id: str):
    """Get import history for a supplier."""
    imports = await db.catalog_imports.find(
        {"supplier_id": supplier_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    return imports


@api_router.get("/catalog/imports/{import_id}/products")
async def get_import_products(import_id: str):
    """Get products from an import."""
    products = await db.catalog_import_products.find(
        {"import_id": import_id},
        {"_id": 0}
    ).to_list(500)
    return products


@api_router.put("/catalog/imports/{import_id}/products/{product_id}")
async def update_import_product(import_id: str, product_id: str):
    """Update a parsed product's data."""
    from fastapi import Request
    # We need the body - let's use a workaround
    pass


class UpdateProductData(BaseModel):
    parsed_data: dict


@api_router.put("/catalog/import-product/{product_id}")
async def update_import_product_data(product_id: str, body: UpdateProductData):
    """Update parsed data for an import product."""
    result = await db.catalog_import_products.update_one(
        {"id": product_id},
        {"$set": {"parsed_data": body.parsed_data}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
    return {"ok": True}


class FinalizeRequest(BaseModel):
    selected_product_ids: List[str]
    supplier_id: str


SALE_UNIT_MAP = {
    "unit": "unit", "unidade": "unit",
    "box": "closed_box", "caixa": "closed_box",
    "bale": "bale", "fardo": "bale",
    "kit": "kit",
    "pair": "pair", "par": "pair",
}


@api_router.post("/catalog/imports/{import_id}/finalize")
async def finalize_import(import_id: str, body: FinalizeRequest):
    """Import selected products to Supabase products table."""
    import_doc = await db.catalog_imports.find_one({"id": import_id})
    if not import_doc:
        raise HTTPException(status_code=404, detail="Importação não encontrada")

    products = await db.catalog_import_products.find(
        {"import_id": import_id, "id": {"$in": body.selected_product_ids}}
    ).to_list(500)

    imported_count = 0
    draft_count = 0
    errors = []

    for prod in products:
        parsed = prod.get("parsed_data", {})
        name = parsed.get("name", "").strip()
        if not name:
            continue

        price = parsed.get("price")
        try:
            price = float(price) if price else 0
        except (ValueError, TypeError):
            price = 0

        min_order = parsed.get("min_order")
        try:
            min_order = int(min_order) if min_order else 1
        except (ValueError, TypeError):
            min_order = 1

        sale_unit_raw = (parsed.get("sale_unit") or "unit").lower().strip()
        sale_unit = SALE_UNIT_MAP.get(sale_unit_raw, "unit")

        variations = parsed.get("variations", [])
        if isinstance(variations, list) and variations:
            variations = [{"name": "Variações", "options": variations}]
        else:
            variations = None

        has_image = bool(parsed.get("image_url"))
        is_draft = not has_image or price <= 0

        images = []
        if parsed.get("image_url"):
            images = [parsed["image_url"]]

        product_data = {
            "id": str(uuid.uuid4()),
            "supplier_id": body.supplier_id,
            "nome": name,
            "descricao_curta": (parsed.get("description") or "")[:500],
            "descricao_longa": parsed.get("description") or "",
            "preco": price,
            "estoque": 0,
            "categoria_id": None,
            "imagens": images,
            "variacoes": variations,
            "brand": parsed.get("brand"),
            "sale_unit": sale_unit,
            "min_order_quantity": min_order,
            "ativo": not is_draft,
            "rating_medio": 0,
            "total_reviews": 0,
            "condition": "new",
            "gender": "none",
            "age_group": "none",
            "keywords": [],
        }

        try:
            result = await supabase_insert_product(product_data)
            new_product_id = result.get("id", product_data["id"])

            await db.catalog_import_products.update_one(
                {"id": prod["id"]},
                {"$set": {"status": "imported", "product_id": new_product_id}}
            )
            imported_count += 1
            if is_draft:
                draft_count += 1
        except Exception as e:
            logger.error(f"Import product error: {e}")
            errors.append({"name": name, "error": str(e)[:100]})

    await db.catalog_imports.update_one(
        {"id": import_id},
        {"$set": {
            "status": "completed",
            "products_imported": imported_count,
            "completed_at": datetime.now(timezone.utc).isoformat(),
        }}
    )

    return {
        "imported": imported_count,
        "drafts": draft_count,
        "errors": errors,
    }


@api_router.get("/catalog/template")
async def download_template():
    """Generate a CSV template for spreadsheet import."""
    from fastapi.responses import StreamingResponse
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Nome", "Descrição", "Preço", "Quantidade Mínima", "Marca", "Categoria", "Variações", "Unidade de Venda"])
    writer.writerow(["Camiseta Básica", "Camiseta 100% algodão", "29.90", "10", "Marca X", "Roupas", "P,M,G,GG", "unit"])
    writer.writerow(["Tênis Esportivo", "Tênis para corrida", "149.90", "5", "Marca Y", "Calçados", "38,39,40,41,42", "pair"])
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=template_nellor.csv"}
    )


# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
